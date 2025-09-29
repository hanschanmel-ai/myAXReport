import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
from tkcalendar import DateEntry
import openpyxl
from openpyxl.utils import get_column_letter
from ax_connector import DynamicsAXConnector

class DynamicsAXReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Dynamics AX 2012 AR Report Generator")
        self.root.geometry("1000x700")
        
        # Initialize AX connector (will be set after login)
        self.ax_connector = None
        self.user_access = None
        
        # Create main frames
        self.create_login_frame()
        self.create_report_frame()
        
        # Show login frame first
        self.login_frame.pack(fill=tk.BOTH, expand=True)
        self.report_frame.pack_forget()
        
    # In the create_login_frame method, remove password field
    def create_login_frame(self):
        self.login_frame = ttk.Frame(self.root, padding=20)
        
        # Login form
        ttk.Label(self.login_frame, text="Dynamics AX 2012 Report Generator", font=("Arial", 16)).pack(pady=10)
        
        form_frame = ttk.Frame(self.login_frame)
        form_frame.pack(pady=20)
        
        ttk.Label(form_frame, text="Server:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.server_entry = ttk.Entry(form_frame, width=30)
        self.server_entry.grid(row=0, column=1, pady=5, padx=5)
        self.server_entry.insert(0, "localhost")
        
        ttk.Label(form_frame, text="Database:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.db_entry = ttk.Entry(form_frame, width=30)
        self.db_entry.grid(row=1, column=1, pady=5, padx=5)
        self.db_entry.insert(0, "DynamicsAX")
        
        ttk.Label(form_frame, text="Username (for role check):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.username_entry = ttk.Entry(form_frame, width=30)
        self.username_entry.grid(row=2, column=1, pady=5, padx=5)
        
        # Remove password field
        
        ttk.Label(form_frame, text="Using Windows Authentication", font=("Arial", 10, "italic")).grid(row=3, column=0, columnspan=2, pady=5)
        
        ttk.Button(self.login_frame, text="Login", command=self.login).pack(pady=10)
    
    # Update the login method
    def login(self):
        server = self.server_entry.get()
        database = self.db_entry.get()
        username = self.username_entry.get()
        
        if not all([server, database, username]):
            messagebox.showerror("Error", "All fields are required")
            return
        
        # Create connector and try to connect using Windows Authentication
        self.ax_connector = DynamicsAXConnector(server, database, username)
        if not self.ax_connector.connect():
            messagebox.showerror("Connection Error", "Failed to connect to Dynamics AX. Please check your server details and ensure you have Windows Authentication access.")
            return
        
        # Get user access level
        self.user_access = self.ax_connector.get_user_access_level(username)
        
        if not self.user_access['has_ar_access']:
            messagebox.showerror("Access Denied", "You do not have access to the Accounts Receivable module.")
            return
        
        # Update UI
        self.user_label.config(text=f"User: {username} | Roles: {', '.join([r['role_name'] for r in self.user_access['roles']])}")
        
        # Switch to report frame
        self.login_frame.pack_forget()
        self.report_frame.pack(fill=tk.BOTH, expand=True)
        
    def create_report_frame(self):
        self.report_frame = ttk.Frame(self.root, padding=20)
        
        # Top bar with user info and logout
        top_bar = ttk.Frame(self.report_frame)
        top_bar.pack(fill=tk.X, pady=(0, 20))
        
        self.user_label = ttk.Label(top_bar, text="User: Not logged in")
        self.user_label.pack(side=tk.LEFT)
        
        ttk.Button(top_bar, text="Logout", command=self.logout).pack(side=tk.RIGHT)
        
        # Report criteria frame
        criteria_frame = ttk.LabelFrame(self.report_frame, text="Report Criteria", padding=10)
        criteria_frame.pack(fill=tk.X, pady=10)
        
        # Create a grid for criteria
        criteria_grid = ttk.Frame(criteria_frame)
        criteria_grid.pack(fill=tk.X, pady=5)
        
        # Customer ID
        ttk.Label(criteria_grid, text="Customer ID:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.customer_id_entry = ttk.Entry(criteria_grid, width=20)
        self.customer_id_entry.grid(row=0, column=1, pady=5, padx=5)
        
        # Date range
        ttk.Label(criteria_grid, text="From Date:").grid(row=0, column=2, sticky=tk.W, pady=5)
        self.from_date_entry = DateEntry(criteria_grid, width=12, background='darkblue',
                                        foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.from_date_entry.grid(row=0, column=3, pady=5, padx=5)
        
        ttk.Label(criteria_grid, text="To Date:").grid(row=0, column=4, sticky=tk.W, pady=5)
        self.to_date_entry = DateEntry(criteria_grid, width=12, background='darkblue',
                                      foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.to_date_entry.grid(row=0, column=5, pady=5, padx=5)
        
        # Invoice ID
        ttk.Label(criteria_grid, text="Invoice ID:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.invoice_id_entry = ttk.Entry(criteria_grid, width=20)
        self.invoice_id_entry.grid(row=1, column=1, pady=5, padx=5)
        
        # Report type
        ttk.Label(criteria_grid, text="Report Type:").grid(row=1, column=2, sticky=tk.W, pady=5)
        self.report_type = tk.StringVar(value="Outstanding Invoices")
        report_types = ["Outstanding Invoices", "Customer Transactions", "Aging Analysis"]
        report_type_combo = ttk.Combobox(criteria_grid, textvariable=self.report_type, values=report_types, state="readonly", width=20)
        report_type_combo.grid(row=1, column=3, pady=5, padx=5)
        
        # Buttons
        button_frame = ttk.Frame(criteria_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Generate Report", command=self.generate_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Criteria", command=self.clear_criteria).pack(side=tk.LEFT, padx=5)
        
        # Results frame
        results_frame = ttk.LabelFrame(self.report_frame, text="Report Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Create Treeview for results
        self.tree = ttk.Treeview(results_frame, columns=("customer_id", "customer_name", "transaction_date", 
                                                        "amount", "due_date", "invoice", "transaction_type"), 
                                show="headings", height=20)
        
        # Define headings
        self.tree.heading("customer_id", text="Customer ID")
        self.tree.heading("customer_name", text="Customer Name")
        self.tree.heading("transaction_date", text="Transaction Date")
        self.tree.heading("amount", text="Amount")
        self.tree.heading("due_date", text="Due Date")
        self.tree.heading("invoice", text="Invoice")
        self.tree.heading("transaction_type", text="Type")
        
        # Define columns
        self.tree.column("customer_id", width=100)
        self.tree.column("customer_name", width=200)
        self.tree.column("transaction_date", width=120)
        self.tree.column("amount", width=100)
        self.tree.column("due_date", width=120)
        self.tree.column("invoice", width=120)
        self.tree.column("transaction_type", width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        # Pack tree and scrollbar
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def login(self):
        server = self.server_entry.get()
        database = self.db_entry.get()
        username = self.username_entry.get()
        
        if not all([server, database, username]):
            messagebox.showerror("Error", "All fields are required")
            return
        
        # Create connector and try to connect
        self.ax_connector = DynamicsAXConnector(server, database, username, password)
        if not self.ax_connector.connect():
            messagebox.showerror("Connection Error", "Failed to connect to Dynamics AX. Please check your credentials.")
            return
        
        # Get user access level
        self.user_access = self.ax_connector.get_user_access_level(username)
        
        if not self.user_access['has_ar_access']:
            messagebox.showerror("Access Denied", "You do not have access to the Accounts Receivable module.")
            return
        
        # Update UI
        self.user_label.config(text=f"User: {username} | Roles: {', '.join([r['role_name'] for r in self.user_access['roles']])}")
        
        # Switch to report frame
        self.login_frame.pack_forget()
        self.report_frame.pack(fill=tk.BOTH, expand=True)
        
    def logout(self):
        if self.ax_connector:
            self.ax_connector.disconnect()
            self.ax_connector = None
            self.user_access = None
        
        # Clear form fields
        self.clear_criteria()
        self.clear_results()
        
        # Switch back to login frame
        self.report_frame.pack_forget()
        self.login_frame.pack(fill=tk.BOTH, expand=True)
        
    def clear_criteria(self):
        self.customer_id_entry.delete(0, tk.END)
        self.invoice_id_entry.delete(0, tk.END)
        today = datetime.date.today()
        self.from_date_entry.set_date(today - datetime.timedelta(days=30))
        self.to_date_entry.set_date(today)
        self.report_type.set("Outstanding Invoices")
        
    def clear_results(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
    def generate_report(self):
        if not self.ax_connector or not self.user_access:
            messagebox.showerror("Error", "Not connected to Dynamics AX")
            return
        
        # Get criteria
        criteria = {
            'customer_id': self.customer_id_entry.get(),
            'from_date': self.from_date_entry.get_date(),
            'to_date': self.to_date_entry.get_date(),
            'invoice_id': self.invoice_id_entry.get(),
            'report_type': self.report_type.get()
        }
        
        # Clear previous results
        self.clear_results()
        
        try:
            # Get data from AX
            ar_data = self.ax_connector.get_ar_report_data(criteria)
            
            # Filter based on report type
            if criteria['report_type'] == "Outstanding Invoices":
                ar_data = [item for item in ar_data if item['due_date'] < datetime.date.today() and item['amount'] > 0]
            elif criteria['report_type'] == "Aging Analysis":
                today = datetime.date.today()
                for item in ar_data:
                    if item['due_date'] < today:
                        days_overdue = (today - item['due_date']).days
                        if days_overdue <= 30:
                            item['aging_bucket'] = "1-30 days"
                        elif days_overdue <= 60:
                            item['aging_bucket'] = "31-60 days"
                        elif days_overdue <= 90:
                            item['aging_bucket'] = "61-90 days"
                        else:
                            item['aging_bucket'] = "90+ days"
                    else:
                        item['aging_bucket'] = "Current"
            
            # Display results
            for item in ar_data:
                self.tree.insert("", tk.END, values=(
                    item['customer_id'],
                    item['customer_name'],
                    item['transaction_date'],
                    f"${item['amount']:.2f}",
                    item['due_date'],
                    item['invoice'],
                    item['transaction_type']
                ))
                
            messagebox.showinfo("Report Generated", f"Found {len(ar_data)} records matching your criteria")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
    
    def export_to_excel(self):
        try:
            # Get column headers
            columns = [self.tree.heading(col)['text'] for col in self.tree['columns']]
            
            # Get data from treeview
            data = []
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                data.append(values)
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            for col_idx, header in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)), 15)
            
            # Save file
            filename = f"AR_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("Export Successful", f"Report exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DynamicsAXReportApp(root)
    root.mainloop()